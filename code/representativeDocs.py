"""
Extract representative documents from BERTopic model and export to Excel.
Each topic gets its own sheet with representative docs and metadata.
"""

import os
import pandas as pd
from bertopic import BERTopic
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configuration
project_root = os.path.expanduser("~/Uncivil-Religion-2.0")
folder_name = "Rescraped03_topic5" # could get from CLI input tbh
model_path = os.path.join(project_root, f"bertopicOutput/{folder_name}/bertopic_model")
output_excel = os.path.join(project_root, f"bertopicOutput/{folder_name}/representative_docs.xlsx")

print("=" * 80)
print("📊 EXTRACTING REPRESENTATIVE DOCUMENTS FROM BERTOPIC MODEL")
print("=" * 80)

# Load the saved model
print(f"\n📂 Loading BERTopic model from: {model_path}")
topic_model = BERTopic.load(model_path)
print("✅ Model loaded successfully")

# Get topic information
topic_info = topic_model.get_topic_info()
print(f"\n📈 Found {len(topic_info)} topics (including outliers)")

# Get representative documents for each topic
print("\n🔍 Extracting representative documents...")

all_data = []

for topic_id in topic_info['Topic'].tolist():
    # Get topic details
    topic_row = topic_info[topic_info['Topic'] == topic_id].iloc[0]
    topic_name = topic_row['Name'] if 'Name' in topic_row else f"Topic_{topic_id}"
    topic_count = topic_row['Count']
    
    # Get representative documents for this topic
    rep_docs = topic_model.get_representative_docs(topic_id)
    # rep_docs = topic_model._extract_representative_docs(
    #     topic=topic_id,
    #     documents=topic_model.documents_,
    #     topics=topic_model.topics_,
    #     nr_repr_docs=5) # Use to get 5 docs regardless of what model was trained with; will be slower
    
    if rep_docs:
        # Get top words for context
        topic_words = topic_model.get_topic(topic_id)
        top_words = ", ".join([word for word, _ in topic_words[:10]]) if topic_words else "N/A"
        
        # Add each representative doc as a row
        for idx, doc in enumerate(rep_docs, 1):
            all_data.append({
                'Topic_ID': topic_id,
                'Topic_Name': topic_name,
                'Topic_Size': topic_count,
                'Top_Words': top_words,
                'Rep_Doc_Number': idx,
                'Document_Length': len(doc),
                'Document_Text': doc
            })
    
# Create DataFrame
df = pd.DataFrame(all_data)

print(f"\n✅ Extracted {len(df)} total representative documents across {len(topic_info)} topics")

# Export to Excel with formatting
print(f"\n💾 Saving to Excel: {output_excel}")

# Create Excel writer
with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    # Write summary sheet
    summary_df = df.groupby('Topic_ID').agg({
        'Topic_Name': 'first',
        'Topic_Size': 'first',
        'Top_Words': 'first',
        'Rep_Doc_Number': 'count'
    }).reset_index()
    summary_df.columns = ['Topic_ID', 'Topic_Name', 'Topic_Size', 'Top_Words', 'Num_Rep_Docs']
    summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    # Write all representative docs to one sheet
    df.to_excel(writer, sheet_name='All_Representative_Docs', index=False)

# Format the Excel file
print("🎨 Applying formatting...")
wb = load_workbook(output_excel)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Format header row
    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for idx, column in enumerate(ws.columns, 1):
        column_letter = get_column_letter(idx)
        max_length = 0
        
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        # Set width (with limits)
        adjusted_width = min(max_length + 2, 100)  # Cap at 100
        if column_letter == get_column_letter(ws.max_column):  # Document text column
            adjusted_width = 80  # Fixed width for document text
        
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Wrap text in document column
    if sheet_name != 'Summary':
        doc_col = get_column_letter(ws.max_column)
        for cell in ws[doc_col]:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Add borders around topic groups in All_Representative_Docs sheet
    if sheet_name == 'All_Representative_Docs':
        print("  Adding borders around topic groups...")
        
        # Group rows by Topic_ID
        current_topic = None
        topic_start_row = None
        
        for row_idx in range(2, ws.max_row + 1):  # Start from 2 (after header)
            topic_id_cell = ws[f'A{row_idx}']
            
            if current_topic is None:
                # First topic group
                current_topic = topic_id_cell.value
                topic_start_row = row_idx
            elif topic_id_cell.value != current_topic:
                # New topic started, apply border to previous group
                for col_idx in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    for r in range(topic_start_row, row_idx):
                        cell = ws[f'{col_letter}{r}']
                        
                        # Apply thick border only on top and bottom rows, NO vertical borders
                        if r == topic_start_row and r == row_idx - 1:
                            # Single row group
                            cell.border = Border(
                                left=Side(style=None),
                                right=Side(style=None),
                                top=Side(style='medium', color='000000'),
                                bottom=Side(style='medium', color='000000')
                            )
                        elif r == topic_start_row:
                            # Top row of group - thick top border only
                            cell.border = Border(
                                left=Side(style=None),
                                right=Side(style=None),
                                top=Side(style='medium', color='000000'),
                                bottom=Side(style=None)
                            )
                        elif r == row_idx - 1:
                            # Bottom row of group - thick bottom border only
                            cell.border = Border(
                                left=Side(style=None),
                                right=Side(style=None),
                                top=Side(style=None),
                                bottom=Side(style='medium', color='000000')
                            )
                
                # Start new group
                current_topic = topic_id_cell.value
                topic_start_row = row_idx
        
        # Apply border to the last topic group
        if topic_start_row is not None:
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                for r in range(topic_start_row, ws.max_row + 1):
                    cell = ws[f'{col_letter}{r}']
                    
                    if r == topic_start_row and r == ws.max_row:
                        # Single row group
                        cell.border = Border(
                            left=Side(style=None),
                            right=Side(style=None),
                            top=Side(style='medium', color='000000'),
                            bottom=Side(style='medium', color='000000')
                        )
                    elif r == topic_start_row:
                        # Top row - thick top border only
                        cell.border = Border(
                            left=Side(style=None),
                            right=Side(style=None),
                            top=Side(style='medium', color='000000'),
                            bottom=Side(style=None)
                        )
                    elif r == ws.max_row:
                        # Bottom row - thick bottom border only
                        cell.border = Border(
                            left=Side(style=None),
                            right=Side(style=None),
                            top=Side(style=None),
                            bottom=Side(style='medium', color='000000')
                        )
    
    # Freeze header row
    ws.freeze_panes = 'A2'

wb.save(output_excel)

print("✅ Excel file created and formatted successfully")

# Print summary statistics
print("\n" + "=" * 80)
print("📊 SUMMARY")
print("=" * 80)
print(f"Total topics analyzed: {len(topic_info)}")
print(f"Total representative documents: {len(df)}")
print(f"Average docs per topic: {len(df) / len(topic_info):.1f}")
print(f"Output file: {output_excel}")
print("\n📁 Excel contains:")
print("  • Summary sheet: Overview of all topics")
print("  • All_Representative_Docs: All docs in one sheet")
print("=" * 80)